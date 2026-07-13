# -*- coding: utf-8 -*-
import openpyxl, re, json
from datetime import date, datetime
from collections import defaultdict, Counter

PY = r"E:\AKshare\020900.xlsx"
wb = openpyxl.load_workbook(PY, data_only=True)
ws = wb["Sheet1"]
rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

def num(s):
    if s is None: return None
    m = re.search(r"-?\d+\.?\d*", str(s))
    return float(m.group()) if m else None

txns = []
for r in rows[1:]:
    a = r[0]
    if not isinstance(a, datetime): continue
    if not isinstance(r[1], str) or "通信设备" not in r[1]: continue
    txns.append([a.date(), r[2], r[3], r[4], r[6]])
txns.sort(key=lambda x: x[0])

flows = []          # [date, cash(+in/-out), btype]
shares = 0.0
excluded = 0
day_nav = {}            # 每个交易日重建净值
day_share_delta = defaultdict(float)   # 每个交易日份额变动
for d, bt, app, conf, st in txns:
    if st and "撤单" in str(st):
        excluded += 1; continue
    if "转入" in bt:
        cash = num(app); sh = num(conf)
        if cash is None: continue
        flows.append([d, -cash, "买"])
        if sh is not None:
            shares += sh
            day_share_delta[d] += sh
            day_nav[d] = cash/sh
    elif "回活期宝" in bt or "卖出" in bt:
        cash = num(conf); sh = num(app)
        if cash is None: continue
        flows.append([d, cash, "卖"])
        if sh is not None:
            shares -= sh
            day_share_delta[d] -= sh
            day_nav[d] = cash/sh

print("有效交易:", len(txns), " 排除撤单:", excluded, " 计入现金流:", len(flows))
print("状态分布:", Counter(t[4] for t in txns))

total_buy = sum(-cf for _, cf, _ in flows if cf < 0)
total_sell = sum(cf for _, cf, _ in flows if cf > 0)
print(f"\n累计申购(现金出): {total_buy:.2f}  累计赎回(现金入): {total_sell:.2f}")
print(f"净投入本金: {total_buy-total_sell:.2f}")
print(f"期末剩余份额: {shares:.2f}  | 4904.59/份额 = {4904.59/shares:.4f} (当前隐含净值)")
print(f"利润(期末值-净投入) = {4904.59-(total_buy-total_sell):.2f}")
nav_samples = list(day_nav.values())
print(f"重建净值样本(前3/后3): {[round(x,3) for x in nav_samples[:3]]} ... {[round(x,3) for x in nav_samples[-3:]]}")

# ---------- 1) XIRR ----------
END_DATE = date(2026, 7, 12); END_ASSET = 4904.59
cfs = [(d, cf) for d, cf, _ in flows] + [(END_DATE, END_ASSET)]
d0 = cfs[0][0]
def xnpv(rate):
    return sum(cf * (1+rate)**(-(d-d0).days/365.0) for d, cf in cfs)

# 扫描所有根
def find_roots(rmin=-0.99, rmax=20.0, step=0.005):
    roots=[]; prev=xnpv(rmin); pr=rmin
    i=1
    while True:
        r=rmin+i*step
        if r>rmax: break
        cur=xnpv(r)
        if prev==0 or (prev<0)!=(cur<0):
            lo,hi=pr,r
            for _ in range(100):
                mid=(lo+hi)/2
                if xnpv(mid)==0: break
                if (xnpv(lo)<0)!=(xnpv(mid)<0): hi=mid
                else: lo=mid
            roots.append((lo+hi)/2)
        prev=cur; pr=r; i+=1
    return roots

print("\n===== XIRR(年化) 扫描所有可能根 =====")
roots = find_roots()
print("所有根:", [round(r*100,2) for r in roots])
xirr = roots[0] if roots else None
if xirr is not None:
    print(f"主根 XIRR = {xirr*100:.4f}%  校验XNPV = {xnpv(xirr):.6f}  区间天数 = {(END_DATE-d0).days}")

# ---------- 2) 简单收益(年化) ----------
net = total_buy - total_sell
simple = END_ASSET/net - 1
simple_ann = (1+simple)**(365.0/(END_DATE-d0).days) - 1
print(f"\n===== 简单收益 =====")
print(f"持有期收益 = {simple*100:.2f}%  年化(复利) = {simple_ann*100:.2f}%")

# ---------- 3) TWR 时间加权 ----------
# 按日期累计份额(用 day_share_delta 重建), 净值用 day_nav
shares_running = 0.0
day_end_shares = {}
for d in sorted(day_share_delta.keys()):
    shares_running += day_share_delta[d]
    day_end_shares[d] = shares_running

# 加上今天(估值日)的当前净值
today = END_DATE
day_nav[today] = END_ASSET/shares
day_end_shares[today] = shares

dates = sorted(set(list(day_nav.keys())))
# TWR = Π (NAV_{i+1}/NAV_i) - 1, 相邻交易日之间份额不变
twr = 1.0
link_log = []
for i in range(len(dates)-1):
    n0 = day_nav[dates[i]]; n1 = day_nav[dates[i+1]]
    if n0 and n1:
        twr *= (n1/n0)
        link_log.append((dates[i], dates[i+1], round(n0,4), round(n1,4), round((n1/n0-1)*100,2)))
twr -= 1
twr_ann = (1+twr)**(365.0/(dates[-1]-dates[0]).days) - 1
print(f"\n===== TWR 时间加权收益率 =====")
print(f"区间: {dates[0]} ~ {dates[-1]} ({ (dates[-1]-dates[0]).days } 天)")
print(f"TWR(持有期) = {twr*100:.2f}%  年化 = {twr_ann*100:.2f}%")
print("链接段示例(前4/后4):")
for x in link_log[:4]: print("  ", x)
for x in link_log[-4:]: print("  ", x)

# ---------- 4) 月度现金流 ----------
print(f"\n===== 月度净现金流 =====")
mflow = defaultdict(float)
for d, cash, bt in flows:
    mflow[(d.year,d.month)] += cash
for k in sorted(mflow):
    print(f"  {k[0]}-{k[1]:02d}: {mflow[k]:+.2f}")

json.dump({
    "xirr": xirr, "xirr_roots": roots,
    "total_buy": total_buy, "total_sell": total_sell, "net": net,
    "shares": shares, "profit": 4904.59-net,
    "simple": simple, "simple_ann": simple_ann,
    "twr": twr, "twr_ann": twr_ann,
    "first_date": str(d0), "days": (END_DATE-d0).days,
    "n_excluded": excluded,
    "link_log": link_log,
}, open(r"E:\AKshare\_diag_tmp.json","w",encoding="utf-8"), ensure_ascii=False)
print("\n诊断数据已保存")
