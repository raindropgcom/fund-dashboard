# -*- coding: utf-8 -*-
"""
批量计算 stocklist 中每只基金的「资金占用内在回报率(严格按日 XIRR)」 v2
- 数据源1: stocklist.xlsx  ->  code | 当前市值
- 数据源2: 基金流水/{code}.xlsx -> 每笔买卖(按确切日期)
- 算法:  Σ CF_i /(1+r)^((d_i - d0)/365) = 0 求 r (XIRR, 与 Excel 一致)
- 业务类型覆盖:
    买入: 活期宝转入 / 银行卡支付买入 / 超级转换(转入) -> 现金流出(负)
    卖出: 卖出回活期宝                -> 现金流入(正)
    现金分红                          -> 现金流入(正)
  * 超级转换只有"份额",用该基金当日隐含净值(由邻近买卖 元/份 反推)估值
  * 撤单/失败(支付失败) 一律剔除
- 输出: 汇总 HTML 网页 + 各基金 JSON
"""
import openpyxl, re, json, os
from datetime import date, datetime

ROOT = r"E:\AKshare"
FLOW_DIR = os.path.join(ROOT, "基金流水")
LIST_FILE = os.path.join(ROOT, "stocklist.xlsx")
VAL_DATE = date(2026, 7, 12)

def num(s):
    if s is None:
        return None
    m = re.search(r"-?\d+\.?\d*", str(s))
    return float(m.group()) if m else None

def solve_xirr(cfs):
    d0 = cfs[0][0]
    def xnpv(rate):
        return sum(cf * (1 + rate) ** (-(d - d0).days / 365.0) for d, cf in cfs)
    roots = []
    prev = xnpv(-0.99); pr = -0.99; i = 1
    while True:
        r = -0.99 + i * 0.005
        if r > 50:
            break
        cur = xnpv(r)
        if prev == 0 or (prev < 0) != (cur < 0):
            lo, hi = pr, r
            for _ in range(200):
                mid = (lo + hi) / 2
                if (xnpv(lo) < 0) != (xnpv(mid) < 0):
                    hi = mid
                else:
                    lo = mid
            roots.append((lo + hi) / 2)
        prev = cur; pr = r; i += 1
    return roots, xnpv

def nav_at(d, nav_series):
    """用买卖反推的隐含净值序列, 对日期 d 线性插值; 越界取最近端点."""
    if not nav_series:
        return None
    nav_series = sorted(nav_series, key=lambda x: x[0])
    if d <= nav_series[0][0]:
        return nav_series[0][1]
    if d >= nav_series[-1][0]:
        return nav_series[-1][1]
    for k in range(len(nav_series) - 1):
        d0, n0 = nav_series[k]; d1, n1 = nav_series[k + 1]
        if d0 <= d <= d1:
            if (d1 - d0).days == 0:
                return (n0 + n1) / 2
            frac = (d - d0).days / (d1 - d0).days
            return n0 + frac * (n1 - n0)
    return nav_series[-1][1]

# ---------- 1) 读取 stocklist ----------
wb_list = openpyxl.load_workbook(LIST_FILE, data_only=True)
ws_list = wb_list["Sheet1"]
funds = {}
for r in ws_list.iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    code = str(r[0]).strip()
    amt = float(r[1]) if r[1] is not None else None
    if code and amt is not None:
        funds[code] = {"amount": amt, "name": None}

results = []
for code in funds:
    fpath = os.path.join(FLOW_DIR, code + ".xlsx")
    if not os.path.exists(fpath):
        results.append({"code": code, "error": f"未找到流水文件 {code}.xlsx"})
        continue

    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # 第一遍: 收集买/卖的隐含净值序列 (元/份)
    nav_series = []
    for r in rows:
        if not isinstance(r[0], datetime) or not isinstance(r[2], str):
            continue
        bt, app, conf, st = r[2], r[3], r[4], (r[6] if len(r) > 6 else None)
        if st and ("撤单" in str(st) or "失败" in str(st)):
            continue
        if ("转入" in bt or "买入" in bt):
            yuan = num(app); sh = num(conf)
            if yuan and sh:
                nav_series.append((r[0].date(), yuan / sh))

    # 第二遍: 建账
    ledger = []
    shares = 0.0
    fund_name = None
    skipped = 0
    status_counter = {}
    assumptions = []
    for r in rows:
        if not isinstance(r[0], datetime) or not isinstance(r[2], str):
            continue
        d, bt, app, conf, st = r[0].date(), r[2], r[3], r[4], (r[6] if len(r) > 6 else None)
        status_counter[st] = status_counter.get(st, 0) + 1
        if fund_name is None and isinstance(r[1], str):
            fund_name = r[1]
        if st and ("撤单" in str(st) or "失败" in str(st)):
            skipped += 1
            continue
        if "转入" in bt or "买入" in bt:           # 买入: 申请=元
            cash = num(app)
            if cash is None:
                skipped += 1; continue
            ledger.append([d, "买入", -cash])
            sh = num(conf)
            if sh: shares += sh
        elif "超级转换" in bt:                      # 转入: 只有份额 -> 用隐含净值估值
            sh = num(app) or num(conf)
            nav = nav_at(d, nav_series)
            if sh is None or nav is None:
                skipped += 1; continue
            val = sh * nav
            ledger.append([d, "超级转换(转入)", -round(val, 2)])
            shares += sh
            assumptions.append(f"{d} 超级转换 {sh:.2f}份, 按当日隐含净值 {nav:.4f} 估值 ≈ {val:.2f}元")
        elif "卖出" in bt or "回活期宝" in bt:      # 卖出: 确认=元
            cash = num(conf)
            if cash is None:
                skipped += 1; continue
            ledger.append([d, "卖出", cash])
            sh = num(app)
            if sh: shares -= sh
        elif "现金分红" in bt:                      # 分红: 确认=元, 现金流入
            cash = num(conf)
            if cash is None:
                skipped += 1; continue
            ledger.append([d, "现金分红", cash])
            assumptions.append(f"{d} 现金分红 +{cash:.2f}元 (计入现金流入)")
        else:
            skipped += 1
    ledger.sort(key=lambda x: x[0])

    if not ledger:
        results.append({"code": code, "error": "无有效交易"})
        continue

    cur_value = funds[code]["amount"]
    cfs = [(d, cash) for d, _, cash in ledger] + [(VAL_DATE, cur_value)]
    roots, xnpv = solve_xirr(cfs)
    xirr = roots[0] if roots else None

    total_buy = sum(-c for _, _, c in ledger if c < 0)
    total_sell = sum(c for _, _, c in ledger if c > 0)
    net = total_buy - total_sell
    profit = cur_value - net
    total_days = (VAL_DATE - cfs[0][0]).days
    cap_days = sum(-cash * (VAL_DATE - d).days for d, _, cash in ledger)
    avg_capital = cap_days / total_days if total_days else 0
    simple_holding = (cur_value / net - 1) if net else None
    simple_annual = ((1 + simple_holding) ** (365 / total_days) - 1) if (simple_holding is not None and total_days) else None
    xnpv_check = xnpv(xirr) if xirr is not None else None

    occ = [[str(d), tp, round(cash, 2), (VAL_DATE - d).days, round(-cash * (VAL_DATE - d).days, 1)]
           for d, tp, cash in ledger]
    occ.append([str(VAL_DATE), "当前市值", round(cur_value, 2), 0, 0.0])

    results.append({
        "code": code, "name": fund_name, "amount": cur_value,
        "xirr": xirr, "roots": roots, "xnpv_check": xnpv_check,
        "total_buy": total_buy, "total_sell": total_sell, "net": net,
        "profit": profit, "shares": shares,
        "first_date": str(cfs[0][0]), "val_date": str(VAL_DATE), "days": total_days,
        "avg_capital": avg_capital, "simple_holding": simple_holding,
        "simple_annual": simple_annual, "n_txn": len(ledger), "skipped": skipped,
        "status_counter": status_counter, "assumptions": assumptions,
        "implied_nav": (cur_value / shares) if shares else None, "ledger": occ,
    })

json.dump(results, open(os.path.join(ROOT, "_fund_irr_batch.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=2)

print(f"估值日: {VAL_DATE}    共 {len(results)} 支基金\n")
print(f"{'代码':<9}{'名称':<22}{'当前市值':>12}{'净投入':>12}{'利润':>10}{'平均占用':>10}{'简单收益':>10}{'XIRR(年)':>12}")
for res in results:
    if "error" in res:
        print(f"{res['code']:<9}  ERROR: {res['error']}"); continue
    s = res["simple_holding"]
    print(f"{res['code']:<9}{str(res['name'])[:20]:<22}{res['amount']:>12.2f}{res['net']:>12.2f}"
          f"{res['profit']:>10.2f}{res['avg_capital']:>10.2f}"
          f"{(s*100 if s is not None else 0):>9.2f}%{((res['xirr']*100) if res['xirr'] else 0):>11.2f}%")
    print(f"   校验XNPV={res['xnpv_check']:.6f}  区间{res['first_date']}~{res['val_date']}({res['days']}天)  "
          f"笔数{res['n_txn']} 剔除{res['skipped']} 隐含净值{res['implied_nav']:.4f}")
    for a in res["assumptions"]:
        print("   假设:", a)
print("\n已保存 _fund_irr_batch.json")
