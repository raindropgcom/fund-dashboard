# -*- coding: utf-8 -*-
"""011062 择时分析(修正版): 用交易反推净值, 套用 择时归因框架(TWR vs XIRR) + 买低卖高指标"""
import openpyxl, re, json, os
from datetime import date, datetime

FLOW = r"E:\AKshare\基金流水\011062.xlsx"
LIST_VAL = 15827.10
VAL_DATE = date(2026, 7, 12)

def num(s):
    if s is None: return None
    m = re.search(r"-?\d+\.?\d*", str(s)); return float(m.group()) if m else None

wb = openpyxl.load_workbook(FLOW, data_only=True)
ws = wb["Sheet1"]
rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

# 预扫: 先建隐含净值序列(由所有买入的 元/份 反推), 不受行序影响
nav_series = []
for r in rows:
    if not isinstance(r[0], datetime) or not isinstance(r[2], str):
        continue
    bt, app, conf, st = r[2], r[3], r[4], (r[6] if len(r) > 6 else None)
    if st and ("撤单" in str(st) or "失败" in str(st)):
        continue
    if ("转入" in bt or "买入" in bt) and num(app) and num(conf):
        nav_series.append((r[0].date(), num(app) / num(conf)))

buys, sells = [], []
for r in rows:
    if not isinstance(r[0], datetime) or not isinstance(r[2], str):
        continue
    d, bt, app, conf, st = r[0].date(), r[2], r[3], r[4], (r[6] if len(r) > 6 else None)
    if st and ("撤单" in str(st) or "失败" in str(st)):
        continue
    if "转入" in bt or "买入" in bt:
        y, s = num(app), num(conf)
        if y and s:
            buys.append({"date": d, "amt": y, "shares": s, "nav": y / s})
    elif "超级转换" in bt:
        s = num(app) or num(conf)
        nd = min(nav_series, key=lambda x: abs((x[0] - d).days))[1] if nav_series else None
        buys.append({"date": d, "amt": s * nd, "shares": s, "nav": nd, "conv": True})
    elif "卖出" in bt or "回活期宝" in bt:
        a, s = num(conf), num(app)
        if a and s:
            sells.append({"date": d, "amt": a, "shares": s, "nav": a / s})

# 剩余份额 / 当前隐含净值 (从列表精确算)
remaining = sum(b["shares"] for b in buys) - sum(s["shares"] for s in sells)
cur_nav = LIST_VAL / remaining

# ---- 指标 ----
total_buy = sum(b["amt"] for b in buys)
total_sell = sum(s["amt"] for s in sells)
net = total_buy - total_sell
avg_buy_nav = sum(b["amt"] * b["nav"] for b in buys) / total_buy
avg_sell_nav = sum(s["amt"] * s["nav"] for s in sells) / total_sell
first_buy_nav = min(b["nav"] for b in buys)
first_buy_date = min(b["date"] for b in buys)
all_nav = [b["nav"] for b in buys] + [s["nav"] for s in sells] + [cur_nav]
min_nav, max_nav = min(all_nav), max(all_nav)
mid_nav = (min_nav + max_nav) / 2

# 基金本身回报 (TWR ≈ 首买净值 -> 当前净值, 期间无分红)
twr_total = cur_nav / first_buy_nav - 1
days_total = (VAL_DATE - first_buy_date).days
twr_annual = (1 + twr_total) ** (365 / days_total) - 1

# XIRR (独立验证过唯一根=4.53%)
def solve_xirr(cfs):
    d0 = cfs[0][0]
    def xnpv(rate): return sum(cf * (1 + rate) ** (-(d - d0).days / 365.0) for d, cf in cfs)
    roots = []; prev = xnpv(-0.99); pr = -0.99; i = 1
    while True:
        r = -0.99 + i * 0.005
        if r > 50: break
        cur = xnpv(r)
        if prev == 0 or (prev < 0) != (cur < 0):
            lo, hi = pr, r
            for _ in range(200):
                mid = (lo + hi) / 2
                if (xnpv(lo) < 0) != (xnpv(mid) < 0): hi = mid
                else: lo = mid
            roots.append((lo + hi) / 2)
        prev = cur; pr = r; i += 1
    return roots

cfs = [(b["date"], -b["amt"]) for b in buys] + [(s["date"], s["amt"]) for s in sells] + [(VAL_DATE, LIST_VAL)]
roots = solve_xirr(cfs)
xirr = roots[0]
xnpv_check = sum(cf * (1 + xirr) ** (-(d - cfs[0][0]).days / 365.0) for d, cf in cfs)
xirr_total = (1 + xirr) ** (days_total / 365) - 1

# 择时归因(核心)
timing_total = xirr_total - twr_total
timing_annual = xirr - twr_annual

for s in sells:
    s["vs_cost"] = s["nav"] / avg_buy_nav - 1
    s["hold_days"] = (s["date"] - first_buy_date).days

low_buy_amt = sum(b["amt"] for b in buys if b["nav"] <= mid_nav)
low_buy_ratio = low_buy_amt / total_buy
loss_sells = [s for s in sells if s["nav"] < avg_buy_nav]
loss_sell_amt = sum(s["amt"] for s in loss_sells)

res = {
    "first_buy_date": str(first_buy_date), "val_date": str(VAL_DATE), "days_total": days_total,
    "first_buy_nav": first_buy_nav, "cur_nav": cur_nav, "min_nav": min_nav, "max_nav": max_nav, "mid_nav": mid_nav,
    "avg_buy_nav": avg_buy_nav, "avg_sell_nav": avg_sell_nav, "sell_premium": avg_sell_nav / avg_buy_nav - 1,
    "total_buy": total_buy, "total_sell": total_sell, "net": net, "current_value": LIST_VAL,
    "twr_total": twr_total, "twr_annual": twr_annual,
    "xirr": xirr, "xirr_total": xirr_total, "xnpv_check": xnpv_check,
    "timing_total": timing_total, "timing_annual": timing_annual,
    "low_buy_ratio": low_buy_ratio, "loss_sell_amt": loss_sell_amt, "loss_sell_cnt": len(loss_sells),
    "buys": buys, "sells": sells, "remaining_shares": remaining,
}
json.dump(res, open(r"E:\AKshare\_011062_timing.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2, default=str)

print(f"[校验] XNPV(xirr)={xnpv_check:.6f} (应≈0)  剩余份额={remaining:.2f} 当前隐含NAV={cur_nav:.4f}")
print(f"首买 {first_buy_date} NAV={first_buy_nav:.4f} | 当前NAV={cur_nav:.4f} | 区间[{min_nav:.4f},{max_nav:.4f}] 中枢={mid_nav:.4f}")
print(f"平均买入NAV={avg_buy_nav:.4f} | 平均卖出NAV={avg_sell_nav:.4f} | 卖/买溢价={res['sell_premium']:.2%}")
print(f"基金本身(TWR)总回报={twr_total:.2%}  年化={twr_annual:.2%}")
print(f"你的XIRR 总回报={xirr_total:.2%}  年化={xirr:.2%}")
print(f">>> 择时归因(核心): XIRR - TWR = {timing_total:.2%} (总) / {timing_annual:.2%} (年化)")
print(f"低位买入占比(NAV<=中枢)={low_buy_ratio:.1%} | 亏损赎回笔数={len(loss_sells)} 金额={loss_sell_amt:.2f}")
print("\n每笔赎回 vs 平均成本:")
for s in sells:
    print(f"  {s['date']} 卖NAV={s['nav']:.4f} 金额={s['amt']:.2f} | 相对成本={s['vs_cost']:.2%} | 持有{s['hold_days']}天")
print("\n已保存 _011062_timing.json")
