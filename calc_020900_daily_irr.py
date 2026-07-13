# -*- coding: utf-8 -*-
"""
020900 资金占用内在回报率(严格按日 XIRR)
- 每一笔买/卖都保留确切日期,不做任何月度合并
- XNPV(r)=Σ CF_i /(1+r)^((d_i - d0)/365) = 0 求 r
- 额外输出每笔资金的"占用天数 & 资金×天数",呈现资金加权本质
"""
import openpyxl, re, json
from datetime import date, datetime

TXN_FILE = r"E:\AKshare\020900.xlsx"
VAL_DATE = date(2026, 7, 12)
END_ASSET = 4904.59

def num(s):
    if s is None: return None
    m = re.search(r"-?\d+\.?\d*", str(s))
    return float(m.group()) if m else None

# 1) 逐笔交易 -> 逐日现金流(保留确切日期)
wb = openpyxl.load_workbook(TXN_FILE, data_only=True)
rows = list(wb['Sheet1'].iter_rows(min_row=1, max_row=wb['Sheet1'].max_row, values_only=True))
ledger = []   # [date, type, cash(+in/-out), shares_delta]
shares = 0.0
excluded = 0
for r in rows:
    if not isinstance(r[0], datetime): continue
    if not isinstance(r[1], str) or "通信设备" not in r[1]: continue
    d, bt, app, conf, st = r[0].date(), r[2], r[3], r[4], r[6]
    if st and "撤单" in str(st):
        excluded += 1; continue
    if "转入" in bt:
        cash = num(app); sh = num(conf)
        if cash is None: continue
        ledger.append([d, "买入", -cash, sh])
        if sh: shares += sh
    elif "回活期宝" in bt or "卖出" in bt:
        cash = num(conf); sh = num(app)
        if cash is None: continue
        ledger.append([d, "卖出", cash, -(sh or 0)])
        if sh: shares -= sh
ledger.sort(key=lambda x: x[0])

# 2) 严格按日 XIRR(逐笔日期, 不做合并)
cfs = [(d, cash) for d, _, cash, _ in ledger] + [(VAL_DATE, END_ASSET)]
d0 = cfs[0][0]
def xnpv(rate):
    return sum(cf * (1+rate)**(-(d-d0).days/365.0) for d, cf in cfs)
def solve():
    roots=[]; prev=xnpv(-0.99); pr=-0.99; i=1
    while True:
        r=-0.99+i*0.005
        if r>30: break
        cur=xnpv(r)
        if prev==0 or (prev<0)!=(cur<0):
            lo,hi=pr,r
            for _ in range(200):
                mid=(lo+hi)/2
                if (xnpv(lo)<0)!=(xnpv(mid)<0): hi=mid
                else: lo=mid
            roots.append((lo+hi)/2)
        prev=cur; pr=r; i+=1
    return roots
roots = solve()
xirr = roots[0] if roots else None

# 3) 资金占用天数分析(每笔到估值日)
occ = []   # date,type,cash,days,cash*days
for d, tp, cash, sh in ledger:
    days = (VAL_DATE - d).days
    occ.append([d, tp, cash, days, -cash*days])   # -cash: 买入为正占用
total_buy = sum(-c for _,_,c,_ in ledger if c < 0)
total_sell = sum(c for _,_,c,_ in ledger if c > 0)
net = total_buy - total_sell
profit = END_ASSET - net
total_days = (VAL_DATE - d0).days
cap_days = sum(-cash*(VAL_DATE-d).days for d,_,cash,_ in ledger)  # 元*天
avg_capital = cap_days / total_days
simple = END_ASSET/net - 1

print("="*64)
print("020900 资金占用内在回报率(严格按日 XIRR)")
print("="*64)
print(f"逐笔现金流笔数: {len(ledger)}  排除已撤单: {excluded}  区间: {d0} ~ {VAL_DATE} ({total_days} 天)")
print(f"累计买入: {total_buy:.2f}  累计卖出: {total_sell:.2f}  净投入: {net:.2f}")
print(f"当前现值: {END_ASSET:.2f}  实际利润: {profit:.2f}")
print(f"期末份额: {shares:.2f}  隐含净值: {END_ASSET/shares:.4f}")
print(f"\n所有 XIRR 根: {[round(r*100,4) for r in roots]}")
print(f"资金占用内在回报率(年化 XIRR) = {xirr*100:.4f}%")
print(f"校验 XNPV(xirr) = {xnpv(xirr):.8f}  (应≈0)")
print(f"\n简单收益率(利润/净投入) = {simple*100:.2f}%  (持有期)")
print(f"平均资金占用 ≈ {avg_capital:.2f} 元 (Σ净流出×天数 / 总天数)")
print(f"简单验证: 利润 {profit:.2f} / 平均占用 {avg_capital:.2f} = {profit/avg_capital*100:.2f}% (占用期总收益率)")
print(f"          年化 ≈ {((1+profit/avg_capital)**(365/total_days)-1)*100:.2f}%")

print("\n" + "="*64)
print("逐笔现金流台账(按确切日期)")
print("="*64)
print(f"{'日期':<12}{'类型':<6}{'现金流':>12}{'占用天数':>8}{'资金×天数':>14}")
for d, tp, cash, days, cd in occ:
    print(f"{str(d):<12}{tp:<6}{cash:>12.2f}{days:>8}{cd:>14.1f}")
print(f"{str(VAL_DATE):<12}{'现值':<6}{END_ASSET:>12.2f}{0:>8}{0:>14.1f}")

json.dump({
    "code":"020900","xirr":xirr,"roots":roots,
    "total_buy":total_buy,"total_sell":total_sell,"net":net,
    "cur_value":END_ASSET,"profit":profit,"shares":shares,
    "first_date":str(d0),"val_date":str(VAL_DATE),"days":total_days,
    "avg_capital":avg_capital,"simple":simple,"excluded":excluded,
    "ledger":[[str(d),tp,cash,days,cd] for d,tp,cash,days,cd in occ],
}, open(r"E:\AKshare\_020900_daily.json","w",encoding="utf-8"), ensure_ascii=False, indent=2)
print("\n已保存 _020900_daily.json")
