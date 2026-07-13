# -*- coding: utf-8 -*-
import openpyxl, re, json
from datetime import date, datetime
from collections import Counter

PY = r"E:\AKshare\020900.xlsx"
wb = openpyxl.load_workbook(PY, data_only=True)
ws = wb["Sheet1"]
rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

def num(s):
    if s is None: return None
    m = re.search(r"-?\d+\.?\d*", str(s))
    return float(m.group()) if m else None

txns = []
for r in rows[1:]:
    a = r[0]
    if not isinstance(a, datetime): continue
    if not isinstance(r[1], str) or "通信设备" not in r[1]: continue
    txns.append([a.date(), r[2], r[3], r[4], r[6]])   # date, 类型, 申请, 确认, 状态
txns.sort(key=lambda x: x[0])

print("有效交易行:", len(txns), " 范围:", txns[0][0], "~", txns[-1][0])
print("业务类型:", Counter(t[1] for t in txns))
print("状态分布:", Counter(t[4] for t in txns))

flows = []   # [date, cash, btype]
shares = 0.0
excluded = 0
for d, bt, app, conf, st in txns:
    if st and "撤单" in str(st):
        excluded += 1
        continue
    if "转入" in bt:                       # 买: 现金流出 = 申请数额(元)
        amt = num(app)
        if amt is None: continue
        flows.append([d, -amt, "买"])
        sh = num(conf)
        if sh is not None: shares += sh
    elif "回活期宝" in bt or "卖出" in bt:  # 卖: 现金流入 = 确认数额(元)
        amt = num(conf)
        if amt is None: continue
        flows.append([d, amt, "卖"])
        sh = num(app)
        if sh is not None: shares -= sh

total_buy = sum(-cf for _, cf, _ in flows if cf < 0)
total_sell = sum(cf for _, cf, _ in flows if cf > 0)
print("\n排除已撤单:", excluded, "笔  | 计入现金流:", len(flows), "笔")
print("累计申购(买):", round(total_buy,2), " 累计赎回(卖):", round(total_sell,2))
print("净投入本金:", round(total_buy-total_sell,2))
print("期末剩余份额(可确认):", round(shares,2), "  | 4904.59/份额=", round(4904.59/shares,4), "(应≈当前净值)")

END_DATE = date(2026, 7, 12)
END_ASSET = 4904.59
cfs = [(d, cf) for d, cf, _ in flows] + [(END_DATE, END_ASSET)]
d0 = cfs[0][0]

def xnpv(rate):
    return sum(cf * (1+rate)**(-(d-d0).days/365.0) for d, cf in cfs)

lo, hi = -0.9999, 50.0
f_lo, f_hi = xnpv(lo), xnpv(hi)
if f_lo*f_hi > 0:
    for h in [200, 1000, 5000]:
        f_hi = xnpv(h)
        if f_lo*f_hi <= 0: hi = h; break
for _ in range(400):
    mid = (lo+hi)/2; f = xnpv(mid)
    if abs(f) < 1e-7: break
    if f_lo*f < 0: hi = mid
    else: lo = mid; f_lo = f
xirr = (lo+hi)/2

print("\n===== XIRR(年化) =", round(xirr*100,4), "% =====")
print("校验 XNPV@XIRR =", round(xnpv(xirr),6), " | 首笔->估值日天数:", (END_DATE-d0).days)

json.dump({
    "flows":[[str(d), bt, round(cf,2)] for d, cf, bt in flows],
    "xirr": xirr, "total_buy": total_buy, "total_sell": total_sell,
    "net": total_buy-total_sell, "shares": shares,
    "implied_nav": 4904.59/shares,
    "d0": str(d0), "end": str(END_DATE), "end_asset": END_ASSET,
    "days":(END_DATE-d0).days, "n_excluded": excluded
}, open(r"E:\AKshare\_xirr_tmp.json","w",encoding="utf-8"), ensure_ascii=False)
print("临时数据已保存")
