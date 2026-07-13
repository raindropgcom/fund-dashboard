# -*- coding: utf-8 -*-
"""
批量计算 stocklist 中每只基金的「聚合择时评价」(全基金·全周期层级)
- 解析口径完全复用 calc_fund_irr_batch.py (买/卖/现金分红/超级转换 + 撤单/失败剔除)
- 价指标:
    净投入 / 当前市值 / 剩余份额 / 期末隐含净值
    XIRR(按日资金加权年化)   -- 二分法求 Σ CF/(1+r)^((d-d0)/365)=0, 回代 XNPV≈0
    TWR (时间加权年化)       -- 用每笔交易 金额÷份额 重建每日NAV, 相邻点连乘
    择时归因 = XIRR - TWR
    平均买NAV / 平均卖NAV / 卖买溢价
    低位买入占比 / 亏损赎回笔数(及金额)
- 输出: _fund_timing_batch.json
"""
import openpyxl, re, json, os
from datetime import date, datetime

ROOT = r"E:\AKshare"
FLOW_DIR = os.path.join(ROOT, "基金流水")
LIST_FILE = os.path.join(ROOT, "stocklist.xlsx")
VAL_DATE = date(2026, 7, 12)

def num(s):
    if s is None:
        return None
    m = re.search(r"-?\d+\.?\d*", str(s))
    return float(m.group()) if m else None

def solve_xirr(cfs):
    """返回 (roots列表, xnpv函数). 与 Excel XIRR 一致."""
    d0 = cfs[0][0]
    def xnpv(rate):
        return sum(cf * (1 + rate) ** (-(d - d0).days / 365.0) for d, cf in cfs)
    roots = []
    prev = xnpv(-0.99); pr = -0.99; i = 1
    while True:
        r = -0.99 + i * 0.005
        if r > 50:
            break
        cur = xnpv(r)
        if prev == 0 or (prev < 0) != (cur < 0):
            lo, hi = pr, r
            for _ in range(200):
                mid = (lo + hi) / 2
                if (xnpv(lo) < 0) != (xnpv(mid) < 0):
                    hi = mid
                else:
                    lo = mid
            roots.append((lo + hi) / 2)
        prev = cur; pr = r; i += 1
    return roots, xnpv

def nav_at(d, nav_series):
    """用买卖反推的隐含净值序列, 对日期 d 线性插值; 越界取最近端点."""
    if not nav_series:
        return None
    nav_series = sorted(nav_series, key=lambda x: x[0])
    if d <= nav_series[0][0]:
        return nav_series[0][1]
    if d >= nav_series[-1][0]:
        return nav_series[-1][1]
    for k in range(len(nav_series) - 1):
        d0, n0 = nav_series[k]; d1, n1 = nav_series[k + 1]
        if d0 <= d <= d1:
            if (d1 - d0).days == 0:
                return (n0 + n1) / 2
            frac = (d - d0).days / (d1 - d0).days
            return n0 + frac * (n1 - n0)
    return nav_series[-1][1]

# ---------- 1) 读取 stocklist ----------
wb_list = openpyxl.load_workbook(LIST_FILE, data_only=True)
ws_list = wb_list["Sheet1"]
funds = {}
for r in ws_list.iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    code = str(r[0]).strip()
    amt = float(r[1]) if r[1] is not None else None
    if code and amt is not None:
        funds[code] = {"amount": amt, "name": None}

results = []
for code in funds:
    fpath = os.path.join(FLOW_DIR, code + ".xlsx")
    warn = []
    if not os.path.exists(fpath):
        results.append({"code": code, "error": f"未找到流水文件 {code}.xlsx"})
        continue

    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # 第一遍: 买/转入 的隐含净值序列(元/份), 供超级转换估值
    nav_series = []
    for r in rows:
        if not isinstance(r[0], datetime) or not isinstance(r[2], str):
            continue
        bt, app, conf, st = r[2], r[3], r[4], (r[6] if len(r) > 6 else None)
        if st and ("撤单" in str(st) or "失败" in str(st)):
            continue
        if ("转入" in bt or "买入" in bt) and num(app) and num(conf):
            nav_series.append((r[0].date(), num(app) / num(conf)))

    # 第二遍: 建账 + 收集买/卖明细(用于NAV重建)
    ledger = []
    trades = []
    shares = 0.0
    fund_name = None
    buys, sells = [], []
    div_amt = 0.0
    for r in rows:
        if not isinstance(r[0], datetime) or not isinstance(r[2], str):
            continue
        d, bt, app, conf, st = r[0].date(), r[2], r[3], r[4], (r[6] if len(r) > 6 else None)
        if fund_name is None and isinstance(r[1], str):
            fund_name = r[1]
        if st and ("撤单" in str(st) or "失败" in str(st)):
            continue
        if "转入" in bt or "买入" in bt:
            cash = num(app); sh = num(conf)
            if cash is None or not sh:
                continue
            ledger.append([d, "买入", -cash])
            shares += sh
            nav = cash / sh
            buys.append({"date": d, "amt": cash, "shares": sh, "nav": nav})
            trades.append({"date": str(d), "type": "买入", "amt": cash, "shares": sh, "nav": nav, "kind": "buy"})
        elif "超级转换" in bt:
            sh = num(app) or num(conf)
            nav = nav_at(d, nav_series)
            if sh is None or nav is None:
                warn.append(f"{d} 超级转换缺份额或无法估值, 已剔除")
                continue
            val = sh * nav
            ledger.append([d, "超级转换(转入)", -round(val, 2)])
            shares += sh
            buys.append({"date": d, "amt": round(val, 2), "shares": sh, "nav": nav, "conv": True})
            trades.append({"date": str(d), "type": "超级转换", "amt": round(val, 2), "shares": sh, "nav": nav, "kind": "convert"})
        elif "卖出" in bt or "回活期宝" in bt:
            cash = num(conf); sh = num(app)
            if cash is None or not sh:
                continue
            ledger.append([d, "卖出", cash])
            shares -= sh
            nav = cash / sh
            sells.append({"date": d, "amt": cash, "shares": sh, "nav": nav})
            trades.append({"date": str(d), "type": "卖出", "amt": cash, "shares": sh, "nav": nav, "kind": "sell"})
        elif "现金分红" in bt:
            cash = num(conf) if num(conf) is not None else num(app)
            if cash is None:
                continue
            ledger.append([d, "现金分红", cash])
            div_amt += cash
            trades.append({"date": str(d), "type": "现金分红", "amt": cash, "shares": None, "nav": None, "kind": "div"})
        # 其它(如调增/调减备注等)忽略
    ledger.sort(key=lambda x: x[0])
    trades.sort(key=lambda x: x["date"])

    if not ledger:
        results.append({"code": code, "error": "无有效交易"})
        continue

    cur_value = funds[code]["amount"]
    cfs = [(d, cash) for d, _, cash in ledger] + [(VAL_DATE, cur_value)]
    roots, xnpv = solve_xirr(cfs)
    xirr = roots[0] if roots else None
    xnpv_check = xnpv(xirr) if xirr is not None else None

    # ----- 基础汇总 -----
    total_buy_amt = sum(-c for _, _, c in ledger if c < 0)          # 买入 + 超级转换
    total_sell_cash = sum(c for _, _, c in ledger if c > 0)         # 卖出 + 现金分红
    sell_amt = sum(s["amt"] for s in sells)
    sell_shares = sum(s["shares"] for s in sells)
    buy_shares = sum(b["shares"] for b in buys)
    net = total_buy_amt - total_sell_cash                           # 净投入(含分红为正向)
    profit = cur_value - net

    if shares <= 1e-9:
        warn.append(f"剩余份额={shares:.4f} ≤ 0, 期末隐含净值无意义")
    cur_nav = (cur_value / shares) if shares > 1e-9 else None

    # ----- TWR: 重建每日NAV, 相邻点连乘 -----
    day_nav = {}
    for b in buys:
        day_nav.setdefault(b["date"], []).append(b["nav"])
    for s in sells:
        day_nav.setdefault(s["date"], []).append(s["nav"])
    day_nav_avg = {d: sum(v) / len(v) for d, v in day_nav.items()}
    trade_days = sorted(day_nav_avg.keys())
    twr = None; twr_annual = None; twr_days = None
    if cur_nav is not None and trade_days:
        nav_pts = [(d, day_nav_avg[d]) for d in trade_days] + [(VAL_DATE, cur_nav)]
        prod = 1.0
        for i in range(len(nav_pts) - 1):
            prod *= nav_pts[i + 1][1] / nav_pts[i][1]
        twr = prod - 1
        twr_days = (VAL_DATE - trade_days[0]).days
        twr_annual = (1 + twr) ** (365 / twr_days) - 1 if twr_days else None

    # ----- 择时归因 -----
    timing = (xirr - twr_annual) if (xirr is not None and twr_annual is not None) else None

    # ----- 平均买/卖净值 & 溢价 -----
    avg_buy_nav = (total_buy_amt / buy_shares) if buy_shares else None
    avg_sell_nav = (sell_amt / sell_shares) if sell_shares else None
    sell_premium = (avg_sell_nav / avg_buy_nav - 1) if (avg_buy_nav and avg_sell_nav) else None

    # ----- 低位买入占比: 买入当日NAV < 全部重建净值点均值 -----
    all_points = list(day_nav_avg.values()) + ([cur_nav] if cur_nav is not None else [])
    mean_nav = (sum(all_points) / len(all_points)) if all_points else None
    low_buy_amt = sum(b["amt"] for b in buys if day_nav_avg.get(b["date"], 1e9) < mean_nav) if mean_nav is not None else 0.0
    low_buy_ratio = (low_buy_amt / total_buy_amt) if total_buy_amt else None

    # ----- 亏损赎回: 卖出当日NAV < 平均买净值 -----
    loss_sells = [s for s in sells if avg_buy_nav is not None and day_nav_avg.get(s["date"], 1e9) < avg_buy_nav]
    loss_sell_cnt = len(loss_sells)
    loss_sell_amt = sum(s["amt"] for s in loss_sells)

    # 收敛性
    if xirr is None:
        warn.append("XIRR 无收敛根 (现金流形态异常)")
    elif xnpv_check is not None and abs(xnpv_check) > 1.0:
        warn.append(f"XNPV 回代={xnpv_check:.4f}, 未收敛到≈0")

    first_date = str(cfs[0][0]); last_date = str(cfs[-2][0]) if len(cfs) > 1 else str(cfs[0][0])
    total_days = (VAL_DATE - cfs[0][0]).days

    # ----- 第二部分「按日 XIRR 汇总」所需支撑指标 (与第一部分同源) -----
    cap_days = sum(-cash * (VAL_DATE - d).days for d, _, cash in ledger)
    avg_capital = cap_days / total_days if total_days else 0
    simple_holding = (cur_value / net - 1) if net else None
    simple_annual = ((1 + simple_holding) ** (365 / total_days) - 1) if (simple_holding is not None and total_days) else None
    occ = [[str(d), tp, round(cash, 2), (VAL_DATE - d).days, round(-cash * (VAL_DATE - d).days, 1)]
           for d, tp, cash in ledger]
    occ.append([str(VAL_DATE), "当前市值", round(cur_value, 2), 0, 0.0])

    results.append({
        "code": code, "name": fund_name, "amount": cur_value,
        "net": net, "profit": profit, "shares": shares, "implied_nav": cur_nav,
        "xirr": xirr, "roots": roots, "xnpv_check": xnpv_check,
        "twr": twr, "twr_annual": twr_annual, "twr_days": twr_days,
        "timing": timing,
        "avg_buy_nav": avg_buy_nav, "avg_sell_nav": avg_sell_nav, "sell_premium": sell_premium,
        "low_buy_ratio": low_buy_ratio, "low_buy_amt": low_buy_amt,
        "loss_sell_cnt": loss_sell_cnt, "loss_sell_amt": loss_sell_amt,
        "mean_nav": mean_nav,
        "total_buy_amt": total_buy_amt, "total_sell_cash": total_sell_cash,
        "sell_amt": sell_amt, "div_amt": div_amt,
        "avg_capital": avg_capital, "simple_holding": simple_holding, "simple_annual": simple_annual,
        "ledger": occ, "trades": trades,
        "first_date": first_date, "last_date": last_date, "val_date": str(VAL_DATE),
        "days": total_days, "n_buys": len(buys), "n_sells": len(sells),
        "warnings": warn,
    })

json.dump(results, open(os.path.join(ROOT, "_fund_timing_batch.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=2)

print(f"估值日: {VAL_DATE}    共 {len(results)} 支基金\n")
print(f"{'代码':<9}{'名称':<22}{'XIRR(年)':>11}{'TWR(年)':>11}{'择时归因':>11}{'卖买溢价':>11}{'低位占比':>11}{'亏损赎回':>10}")
for res in results:
    if "error" in res:
        print(f"{res['code']:<9}  ERROR: {res['error']}")
        continue
    def pc(v, d=2):
        return "—" if v is None else f"{v*100:.{d}f}%"
    print(f"{res['code']:<9}{str(res['name'])[:20]:<22}"
          f"{pc(res['xirr']):>11}{pc(res['twr_annual']):>11}{pc(res['timing']):>11}"
          f"{pc(res['sell_premium']):>11}{pc(res['low_buy_ratio']):>11}{res['loss_sell_cnt']:>8}笔")
    print(f"   校验XNPV={res['xnpv_check']:.6f}  净投入={res['net']:.2f} 利润={res['profit']:.2f} "
          f"剩余份额={res['shares']:.2f} 隐含NAV={res['implied_nav']:.4f} 平均买NAV={res['avg_buy_nav']:.4f} 平均卖NAV={res['avg_sell_nav']:.4f}")
    for w in res["warnings"]:
        print("   ⚠", w)
print("\n已保存 _fund_timing_batch.json")
