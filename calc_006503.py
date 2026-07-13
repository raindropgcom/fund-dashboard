# -*- coding: utf-8 -*-
import openpyxl, re, json
from datetime import date, datetime
from collections import defaultdict, Counter

TXN_FILE = r"E:\AKshare\006503.xlsx"
LIST_FILE = r"E:\AKshare\stocklist.xlsx"
VAL_DATE = date(2026, 7, 12)   # 估值日(今天)

# ---- 1) 从 stocklist 取 006503 当前现值 ----
wb2 = openpyxl.load_workbook(LIST_FILE, data_only=True)
ws2 = wb2['Sheet1']
cur_value = None
for r in ws2.iter_rows(min_row=2, values_only=True):
    if r[0] is None: continue
    code = str(r[0]).strip()
    if code in ("006503", "6503"):
        cur_value = float(r[1]); break
print("006503 当前现值(来自 stocklist):", cur_value)

# ---- 2) 解析交易流水 ----
wb = openpyxl.load_workbook(TXN_FILE, data_only=True)
ws = wb['Sheet1']
rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

def num(s):
    if s is None: return None
    m = re.search(r"-?\d+\.?\d*", str(s))
    return float(m.group()) if m else None

txns = []
for r in rows:
    if not isinstance(r[0], datetime): continue       # 跳过 noise 行(datetime.time/其它)
    if not isinstance(r[2], str): continue
    txns.append([r[0].date(), r[2], r[3], r[4], r[6]])
txns.sort(key=lambda x: x[0])

flows = []            # [date, cash(+in/-out), btype]
shares = 0.0
day_nav = {}
day_share_delta = defaultdict(float)
excluded = 0
for d, bt, app, conf, st in txns:
    if st and "撤单" in str(st):
        excluded += 1; continue
    if "转入" in bt:                                # 买: 申请=元(流出), 确认=份
        cash = num(app); sh = num(conf)
        if cash is None: continue
        flows.append([d, -cash, "买"])
        if sh is not None:
            shares += sh; day_share_delta[d] += sh; day_nav[d] = cash/sh
    elif "回活期宝" in bt or "卖出" in bt:          # 卖: 申请=份, 确认=元(流入)
        cash = num(conf); sh = num(app)
        if cash is None: continue
        flows.append([d, cash, "卖"])
        if sh is not None:
            shares -= sh; day_share_delta[d] -= sh; day_nav[d] = cash/sh

print("有效交易:", len(txns), " 排除撤单:", excluded, " 计入现金流:", len(flows))
print("状态分布:", Counter(t[4] for t in txns))
total_buy = sum(-cf for _, cf, _ in flows if cf < 0)
total_sell = sum(cf for _, cf, _ in flows if cf > 0)
print(f"\n累计申购(现金出): {total_buy:.2f}  累计赎回(现金入): {total_sell:.2f}")
print(f"净投入本金: {total_buy-total_sell:.2f}")
print(f"期末剩余份额: {shares:.2f}  | {cur_value}/份额 = {cur_value/shares:.4f} (当前隐含净值)")
print(f"利润(现值-净投入) = {cur_value-(total_buy-total_sell):.2f}")
nav_s = list(day_nav.values())
print(f"重建净值样本(首3/末3): {[round(x,3) for x in nav_s[:3]]} ... {[round(x,3) for x in nav_s[-3:]]}")

# ---- 3) XIRR ----
END_ASSET = cur_value
cfs = [(d, cf) for d, cf, _ in flows] + [(VAL_DATE, END_ASSET)]
d0 = cfs[0][0]
def xnpv(rate):
    return sum(cf * (1+rate)**(-(d-d0).days/365.0) for d, cf in cfs)

def find_roots(rmin=-0.99, rmax=20.0, step=0.005):
    roots=[]; prev=xnpv(rmin); pr=rmin; i=1
    while True:
        r=rmin+i*step
        if r>rmax: break
        cur=xnpv(r)
        if prev==0 or (prev<0)!=(cur<0):
            lo,hi=pr,r
            for _ in range(100):
                mid=(lo+hi)/2
                if xnpv(mid)==0: break
                if (xnpv(lo)<0)!=(xnpv(mid)<0): hi=mid
                else: lo=mid
            roots.append((lo+hi)/2)
        prev=cur; pr=r; i+=1
    return roots

roots = find_roots()
print("\n===== XIRR(年化) 所有根 =====", [round(r*100,2) for r in roots])
xirr = roots[0] if roots else None
if xirr is not None:
    print(f"XIRR = {xirr*100:.4f}%  校验XNPV = {xnpv(xirr):.6f}  区间天数 = {(VAL_DATE-d0).days}")

# ---- 4) 简单收益 ----
net = total_buy - total_sell
simple = END_ASSET/net - 1
simple_ann = (1+simple)**(365.0/(VAL_DATE-d0).days) - 1
print(f"\n===== 简单收益 ===== 持有期 {simple*100:.2f}%  年化 {simple_ann*100:.2f}%")

# ---- 5) TWR ----
shares_running = 0.0
day_end_shares = {}
for d in sorted(day_share_delta.keys()):
    shares_running += day_share_delta[d]
    day_end_shares[d] = shares_running
today = VAL_DATE
day_nav[today] = END_ASSET/shares
day_end_shares[today] = shares
dates = sorted(set(day_nav.keys()))
twr = 1.0; link_log = []
for i in range(len(dates)-1):
    n0 = day_nav[dates[i]]; n1 = day_nav[dates[i+1]]
    if n0 and n1:
        twr *= (n1/n0)
        link_log.append((dates[i], dates[i+1], round(n0,4), round(n1,4), round((n1/n0-1)*100,2)))
twr -= 1
twr_ann = (1+twr)**(365.0/(dates[-1]-dates[0]).days) - 1
print(f"\n===== TWR 时间加权 ===== 区间 {dates[0]}~{dates[-1]} ({ (dates[-1]-dates[0]).days }天)")
print(f"TWR(持有期) = {twr*100:.2f}%  年化 = {twr_ann*100:.2f}%")

# ---- 6) 月度现金流 ----
print(f"\n===== 月度净现金流 =====")
mflow = defaultdict(float)
for d, cash, bt in flows:
    mflow[(d.year,d.month)] += cash
for k in sorted(mflow):
    print(f"  {k[0]}-{k[1]:02d}: {mflow[k]:+.2f}")

json.dump({
    "code":"006503","xirr":xirr,"xirr_roots":roots,
    "total_buy":total_buy,"total_sell":total_sell,"net":net,
    "shares":shares,"cur_value":cur_value,"profit":cur_value-net,
    "simple":simple,"simple_ann":simple_ann,
    "twr":twr,"twr_ann":twr_ann,
    "first_date":str(d0),"days":(VAL_DATE-d0).days,"n_excluded":excluded,
    "link_log":[(str(a),str(b),c,d,e) for a,b,c,d,e in link_log],
}, open(r"E:\AKshare\_006503_tmp.json","w",encoding="utf-8"), ensure_ascii=False)
print("\n计算完成，数据已保存")
