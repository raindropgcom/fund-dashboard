# -*- coding: utf-8 -*-
import openpyxl
from datetime import datetime, date
import numpy as np

PY = r"E:\AKshare\020900.xlsx"
wb = openpyxl.load_workbook(PY, data_only=True)
ws = wb["Sheet1"]
rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

txns = []
for r in rows:
    d, name, btype, apply_amt, conf_amt, acct, status, op, nav = r[:9]
    nav = float(nav)
    if btype == "买":
        cash = -float(apply_amt); shares = float(conf_amt)
    else:
        cash = float(conf_amt); shares = -float(apply_amt)
    txns.append([d.date(), btype, cash, shares, nav])
txns.sort(key=lambda x: x[0])

END_DATE = date(2026, 7, 12)
END_ASSET = 4904.59
LAST_NAV = txns[-1][4]

# 份额与现金流汇总
shares = 0.0; cum_buy = 0.0; cum_sell = 0.0
flow_rows = []
for d, bt, cash, sh, nav in txns:
    shares += sh
    if cash < 0: cum_buy += -cash
    else: cum_sell += cash
    flow_rows.append((d, bt, cash, sh, nav, shares))
final_shares = shares
value_at_last_nav = final_shares * LAST_NAV
implied_nav_now = END_ASSET / final_shares

def xnpv(rate, cfs):
    d0 = cfs[0][0]
    return sum(cf * (1+rate)**(-(d-d0).days/365.0) for d, cf in cfs)

def xirr(cfs):
    lo, hi = -0.9999, 50.0
    f_lo, f_hi = xnpv(lo, cfs), xnpv(hi, cfs)
    if f_lo*f_hi > 0:
        for h in [200, 1000, 5000]:
            f_hi = xnpv(h, cfs)
            if f_lo*f_hi <= 0: hi = h; break
        else: return None
    for _ in range(300):
        mid=(lo+hi)/2; f=xnpv(mid, cfs)
        if abs(f)<1e-9: return mid
        if f_lo*f<0: hi=mid
        else: lo=mid; f_lo=f
    return (lo+hi)/2

cfs_base = [(d, cash) for d, bt, cash, sh, nav in txns]

# 敏感度表: 不同期末资产对应的 XIRR
sens = []
for fv in [round(value_at_last_nav,2), 500, 1000, 2000, 3000, END_ASSET]:
    cfs = cfs_base + [(END_DATE, fv)]
    r = xirr(cfs)
    sens.append((fv, r))

# 用用户给定值
xirr_user = xirr(cfs_base + [(END_DATE, END_ASSET)])

print("期末剩余份额:", round(final_shares,4))
print("按末笔交易日净值估算持仓市值:", round(value_at_last_nav,2))
print("给定4904.59倒推当前净值:", round(implied_nav_now,4), "(交易期净值仅2.6~2.75, 矛盾!)")
print("累计买入:", round(cum_buy,2), " 累计卖出:", round(cum_sell,2), " 净投入:", round(cum_buy-cum_sell,2))
print("XIRR(用户给定4904.59):", None if xirr_user is None else round(xirr_user*100,2),"%")
print("敏感度(期末资产 -> XIRR年化):")
for fv,r in sens:
    print(f"  {fv:>10.2f} -> {('N/A' if r is None else str(round(r*100,2))+'%')}")

# ---- 生成 HTML 报告 ----
html = []
html.append("""<html><head><meta charset='utf-8'><style>
body{font-family:-apple-system,'Segoe UI','Microsoft YaHei',sans-serif;background:#f7f8fa;color:#1f2329;max-width:1000px;margin:24px auto;padding:0 18px;}
h1{font-size:22px;border-left:5px solid #2f6fed;padding-left:10px;}
h2{font-size:17px;margin-top:30px;color:#2f6fed;}
table{border-collapse:collapse;width:100%;margin:12px 0;font-size:13px;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.08);}
th,td{border:1px solid #e5e6eb;padding:6px 8px;text-align:center;}
th{background:#eef2ff;}
.out{color:#d4380d;font-weight:bold;} .in{color:#389e0d;font-weight:bold;}
.warn{background:#fff7e6;border:1px solid #ffd591;border-radius:8px;padding:14px 16px;margin:16px 0;color:#874d00;}
.calc{background:#e6fffb;border:1px solid #87e8de;border-radius:8px;padding:14px 16px;margin:16px 0;}
.kpi{display:flex;gap:14px;flex-wrap:wrap;margin:14px 0;}
.kpi div{background:#fff;border:1px solid #e5e6eb;border-radius:8px;padding:12px 16px;flex:1;min-width:150px;}
.kpi b{display:block;font-size:20px;color:#2f6fed;}
.method{background:#fff;border:1px solid #e5e6eb;border-radius:8px;padding:12px 16px;margin:10px 0;}
code{background:#f0f2f5;padding:1px 5px;border-radius:4px;}
</style></head><body>""")
html.append("<h1>基金 020900 收益率（IRR）分析报告</h1>")
html.append(f"<p>产品：<b>天弘中证全指通信设备指数发起C</b> ｜ 交易区间：{txns[0][0]} ~ {txns[-1][0]} ｜ 估值日：{END_DATE}</p>")

html.append("<div class='warn'><b>⚠️ 数据矛盾提示（务必先看）</b><br>")
html.append(f"根据流水，期末您持有 <b>{final_shares:.2f}</b> 份。交易期单位净值在 <b>2.59~2.75</b> 之间，")
html.append(f"因此按末笔净值估算持仓市值约为 <b>{value_at_last_nav:.2f} 元</b>。")
html.append(f"但您给出的「当前基金资产 4904.59 元」意味着当前净值高达 <b>{implied_nav_now:.2f}</b> —— 是交易期的约 <b>{implied_nav_now/LAST_NAV:.1f} 倍</b>，")
html.append("在 4.5 个月内对指数基金几乎不可能。请先确认：")
html.append("<ul><li>4904.59 是否其实是<b>整个账户</b>的总资产（含其它基金）？</li>")
html.append("<li>或 020900 在 2026-01-19 <b>之前已有持仓</b>（本文件未包含），需要先确认初始份额与成本？</li>")
html.append("<li>或该数字有误 / 单位不同？</li></ul>")
html.append("下面在「按您给定 4904.59 计算」和「按真实持仓推算」两种口径下分别给出结果。</div>")

html.append("<h2>一、现金流梳理（已核对）</h2>")
html.append("<div class='kpi'>")
html.append(f"<div>累计申购(买)<b>{cum_buy:.2f}元</b></div>")
html.append(f"<div>累计赎回(卖)<b>{cum_sell:.2f}元</b></div>")
html.append(f"<div>净投入本金<b>{cum_buy-cum_sell:.2f}元</b></div>")
html.append(f"<div>期末份额<b>{final_shares:.2f}份</b></div>")
html.append("</div>")
html.append("<table><tr><th>日期</th><th>类型</th><th>现金流(元)</th><th>份额变动</th><th>当日净值</th><th>累计份额</th></tr>")
for d, bt, cash, sh, nav, shc in flow_rows:
    cls = "out" if cash<0 else "in"
    sign = "-" if cash<0 else "+"
    html.append(f"<tr><td>{d}</td><td>{bt}</td><td class='{cls}'>{sign}{abs(cash):.2f}</td><td>{sh:+.2f}</td><td>{nav:.4f}</td><td>{shc:.2f}</td></tr>")
html.append(f"<tr><td>{END_DATE}</td><td>期末清算(假设)</td><td class='in'>+{END_ASSET:.2f}</td><td></td><td>?</td><td></td></tr>")
html.append("</table>")

html.append("<h2>二、用您给定的 4904.59 计算（资金加权 XIRR）</h2>")
if xirr_user is not None:
    html.append(f"<div class='calc'>XIRR（年化内部收益率）≈ <b>{xirr_user*100:,.0f}%</b><br>")
    html.append("<small>这个值在数学上成立，但明显是上面「数据矛盾」造成的假象——它等价于要求基金在 ~5 个月内涨 20 多倍。请勿直接使用。</small></div>")
else:
    html.append("<div class='calc'>无实数解</div>")

html.append("<h2>三、敏感度：期末资产取不同值时对应的 XIRR</h2>")
html.append("<p>下表说明结果对「期末资产」极其敏感。若 4904.59 其实是账户总值，可据此反推 020900 占多少。</p>")
html.append("<table><tr><th>假设的期末基金资产(元)</th><th>XIRR 年化</th><th>说明</th></tr>")
for fv,r in sens:
    note = ""
    if abs(fv-value_at_last_nav)<1: note="按末笔交易日净值估算的真实持仓市值"
    if abs(fv-END_ASSET)<1: note="用户给定值（含矛盾）"
    rr = "N/A" if r is None else f"{r*100:,.1f}%"
    html.append(f"<tr><td>{fv:.2f}</td><td><b>{rr}</b></td><td>{note}</td></tr>")
html.append("</table>")

html.append("<h2>四、计算 IRR 的 4 种方法（给你系统介绍）</h2>")
methods = [
("1. XIRR（资金加权·按实际日期）— 最推荐","Excel 的 XIRR。把每一笔现金流按<b>真实发生日期</b>折现，求使净现值=0 的折现率。公式：Σ CFᵢ/(1+r)^((dᵢ−d₀)/365)=0。",
 "适合现金流<b>不定期、金额不等</b>的真实投资（你这笔就是）。它考虑了你的买卖时点，所以也叫「钱加权收益率」——收益率高低会被你的择时放大或缩小。"),
("2. IRR（资金加权·假设等间隔）","Excel 的 IRR。假设每笔现金流间隔相等（如每月一期），求 Σ CFₜ/(1+r)^t=0。",
 "只适合<b>周期固定</b>的场景（如每月定投）。你的交易间隔乱，直接套 IRR 会失真，本例按序算得约 19.7%/期，仅作对照，不能用。"),
("3. TWR 时间加权收益率（剔除择时）","在每笔交易日按当日净值给组合「估值」，把相邻两次估值之间的收益率连乘：TWR=Π(1+rₜ)−1。",
 "它<b>剔除你买卖时点的影响</b>，纯粹反映基金本身涨跌，是评价基金经理/基金表现的行业金标准。你拿它跟指数比，才知道基金赚不赚钱。"),
("4. 修正 Dietz 法（TWR 的简便近似）","用 (期末值−净现金流)/(期初值+按日期加权的净现金流) 近似单段收益，可分段做。",
 "比完整 TWR 省事（不必每段重新估值），精度足够日常用，很多基金账户App的「收益率」就是它。"),
]
for t,d,u in methods:
    html.append(f"<div class='method'><b>{t}</b><br><span>{d}</span><br><small><b>怎么用：</b>{u}</small></div>")

html.append("<h2>五、建议的下一步</h2>")
html.append("<ul>")
html.append("<li>确认 4904.59 到底是 020900 的单基市值，还是账户总市值。</li>")
html.append("<li>若是单基且确实有 020900 的早期持仓，请提供 <b>2026-01-19 之前的初始份额与成本</b>，我才能算对 IRR（IRR 必须包含全部本金投入）。</li>")
html.append("<li>若只有本文件这些交易、且当前真值约 216 元，告诉我当前准确净值或份额，我立刻给出可采信的 XIRR 与 TWR。</li>")
html.append("</ul>")
html.append(f"<p style='color:#8a8f99;font-size:12px;'>* 报告生成于 {datetime.now():%Y-%m-%d %H:%M} ｜ 估值日 {END_DATE}</p>")
html.append("</body></html>")

with open(r"E:\AKshare\IRR报告_020900.html","w",encoding="utf-8") as f:
    f.write("\n".join(html))
print("\nHTML report written: E:\\AKshare\\IRR报告_020900.html")
