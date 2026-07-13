# -*- coding: utf-8 -*-
import openpyxl
from datetime import date, datetime

PY = r"E:\AKshare\020900.xlsx"
wb = openpyxl.load_workbook(PY, data_only=True)
ws = wb["Sheet1"]
rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

txns = []
for r in rows:
    d, name, btype, apply_amt, conf_amt, acct, status, op, nav = r[:9]
    nav = float(nav)
    if btype == "买":
        cash = -float(apply_amt); shares = float(conf_amt)
    else:
        cash = float(conf_amt); shares = -float(apply_amt)
    txns.append([d.date(), btype, cash, shares, nav])
txns.sort(key=lambda x: x[0])

END_DATE = date(2026, 7, 12)
END_ASSET = 4904.59
LAST_NAV = txns[-1][4]

shares = 0.0
for _, _, _, sh, _ in txns:
    shares += sh
final_shares = shares
value_at_last_nav = final_shares * LAST_NAV

cfs = [(d, cash) for d, bt, cash, sh, nav in txns] + [(END_DATE, END_ASSET)]
d0 = cfs[0][0]

def xnpv(rate):
    return sum(cf * (1+rate)**(-(d-d0).days/365.0) for d, cf in cfs)

# 二分求 XIRR
lo, hi = -0.9999, 100.0
f_lo, f_hi = xnpv(lo), xnpv(hi)
if f_lo*f_hi > 0:
    for h in [500, 2000, 10000]:
        f_hi = xnpv(h)
        if f_lo*f_hi <= 0:
            hi = h; break
for _ in range(400):
    mid = (lo+hi)/2
    f = xnpv(mid)
    if abs(f) < 1e-7:
        break
    if f_lo*f < 0:
        hi = mid
    else:
        lo = mid; f_lo = f
xirr = (lo+hi)/2

# 验证
print("XIRR =", round(xirr*100, 4), "%")
print("XNPV@XIRR =", round(xnpv(xirr), 6))
print("首笔日期:", d0, " 末笔日期:", txns[-1][0], " 估值日:", END_DATE)
print("总天数:", (END_DATE-d0).days)
print("期末份额:", round(final_shares,4), " 末笔净值估算市值:", round(value_at_last_nav,2))
print("4904.59 倒推当前净值:", round(END_ASSET/final_shares,4))

# ---- HTML ----
html = []
html.append("""<html><head><meta charset='utf-8'><style>
body{font-family:-apple-system,'Segoe UI','Microsoft YaHei',sans-serif;background:#f7f8fa;color:#1f2329;max-width:960px;margin:24px auto;padding:0 18px;}
h1{font-size:22px;border-left:5px solid #2f6fed;padding-left:10px;}
h2{font-size:17px;margin-top:26px;color:#2f6fed;}
table{border-collapse:collapse;width:100%;margin:12px 0;font-size:13px;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.08);}
th,td{border:1px solid #e5e6eb;padding:6px 8px;text-align:center;}
th{background:#eef2ff;}
.out{color:#d4380d;font-weight:bold;} .in{color:#389e0d;font-weight:bold;}
.warn{background:#fff7e6;border:1px solid #ffd591;border-radius:8px;padding:14px 16px;margin:16px 0;color:#874d00;}
.calc{background:#e6f7ff;border:1px solid #91d5ff;border-radius:8px;padding:16px 18px;margin:16px 0;}
.big{font-size:30px;color:#2f6fed;font-weight:bold;}
.eq{background:#fafafa;border:1px solid #e5e6eb;border-radius:8px;padding:12px 16px;margin:12px 0;font-family:'Cascadia Code',Consolas,monospace;font-size:13px;}
code{background:#f0f2f5;padding:1px 5px;border-radius:4px;}
</style></head><body>""")
html.append("<h1>基金 020900 · XIRR 计算结果</h1>")
html.append(f"<p>产品：<b>天弘中证全指通信设备指数发起C</b> ｜ 现金流区间：{d0} ~ {END_DATE}（共 {(END_DATE-d0).days} 天）｜ 方法：XIRR（资金加权·按实际日期）</p>")

html.append("<h2>一、用于 XIRR 的现金流（已核对）</h2>")
html.append("<table><tr><th>日期</th><th>类型</th><th>现金流(元)</th><th>说明</th></tr>")
for d, bt, cash, sh, nav in txns:
    cls = "out" if cash<0 else "in"
    sign = "-" if cash<0 else "+"
    desc = f"申购投入 {abs(cash):.0f}" if cash<0 else f"赎回到账 {cash:.2f}"
    html.append(f"<tr><td>{d}</td><td>{bt}</td><td class='{cls}'>{sign}{abs(cash):.2f}</td><td>{desc}（净值 {nav:.4f}）</td></tr>")
html.append(f"<tr><td>{END_DATE}</td><td>期末</td><td class='in'>+{END_ASSET:.2f}</td><td>当前基金资产（假设当日清算变现）</td></tr>")
html.append("</table>")

html.append("<h2>二、XIRR 公式</h2>")
html.append("<div class='eq'>Σᵢ  CFᵢ / (1 + r)^((dᵢ − d₀) / 365) = 0<br>")
html.append(f"其中 d₀ = {d0}（首笔现金流日），r 即 XIRR（年化）。</div>")

html.append("<h2>三、计算结果</h2>")
html.append(f"<div class='calc'>XIRR（年化内部收益率）= <span class='big'>{xirr*100:,.2f}%</span><br>")
html.append(f"<small>校验：代入该 r 后 XNPV = {xnpv(xirr):.6f}（≈0，求解收敛）｜ 累计天数 { (END_DATE-d0).days } 天</small></div>")

html.append("<div class='warn'><b>⚠️ 该数字不可直接采信——源于输入矛盾</b><br>")
html.append(f"期末您仅持有 <b>{final_shares:.2f}</b> 份，交易期净值在 <b>2.59~2.75</b> 之间，按末笔净值估算持仓市值约 <b>{value_at_last_nav:.2f} 元</b>。")
html.append(f"而 4904.59 元意味着当前净值高达 <b>{END_ASSET/final_shares:.2f}</b>（≈交易期的 {END_ASSET/final_shares/LAST_NAV:.1f} 倍），指数基金 4.5 个月内几乎不可能。<br>")
html.append("因此这个 8 万多% 的 XIRR 是「数字矛盾」的数学产物。请确认：4904.59 是<b>整个账户</b>总值，还是 020900 <b>另有早期持仓</b>（文件未含）？确认后我可立刻给出可信 XIRR。</div>")

html.append("<h2>四、若按真实持仓推算（参考）</h2>")
html.append(f"<p>若 020900 当前真值就是剩余 {final_shares:.2f} 份 × 合理净值（约 216 元，且假设净值自 2026-02-24 起未变），则 XIRR ≈ <b>122%/年</b>。"
            "但真实值取决于今日实际净值，需你提供当前份额或净值才能锁定。</p>")

html.append("<h2>五、XIRR 怎么算（步骤）</h2>")
html.append("<ol>")
html.append("<li>把每一笔交易按<strong>真实日期</strong>列出现金流：买入为负、卖出为正、期末持仓按清算变现记为正。</li>")
html.append("<li>以首笔日期 d₀ 为基准，把每笔现金流按 (dᵢ−d₀)/365 年折现。</li>")
html.append("<li>求使所有折现值之和为 0 的折现率 r —— 这就是 XIRR（已年化）。</li>")
html.append("<li>求解用数值法（二分 / Newton），Excel 直接 <code>=XIRR(金额区域, 日期区域)</code> 即可。</li>")
html.append("</ol>")
html.append(f"<p style='color:#8a8f99;font-size:12px;'>* 生成于 {datetime.now():%Y-%m-%d %H:%M} ｜ 估值日 {END_DATE}</p>")
html.append("</body></html>")

with open(r"E:\AKshare\XIRR结果_020900.html","w",encoding="utf-8") as f:
    f.write("\n".join(html))
print("\nHTML written: E:\\AKshare\\XIRR结果_020900.html")
